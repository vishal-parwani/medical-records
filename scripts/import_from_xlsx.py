"""One-time import: Family_Medical_Record.xlsx -> Firestore.

Usage:
    pip install openpyxl firebase-admin
    export GOOGLE_APPLICATION_CREDENTIALS=/path/to/serviceAccount.json
    python3 scripts/import_from_xlsx.py Family_Medical_Record.xlsx --dry-run
    python3 scripts/import_from_xlsx.py Family_Medical_Record.xlsx

Run with --dry-run first and check the printed counts/samples before writing.
This is meant to run ONCE against an empty database — re-running appends
duplicate records (each row becomes a new auto-ID doc), it does not upsert.

Each per-member sheet is a loose, human-maintained document: section titles,
column headers and even layout (e.g. personal info as one label/value pair
per row vs. two pairs per row) vary between sheets. This parser detects
section boundaries by the "  SECTION NAME" marker rows (two leading spaces)
used consistently in this workbook, and maps header cell text to canonical
field keys by keyword match rather than exact position, so it tolerates the
inconsistencies found sheet-to-sheet. Rows/sections that don't match a known
pattern are skipped and reported at the end for a manual look — spot-check
that list against the original xlsx after import.
"""
import argparse
import re
import sys

import openpyxl

MEMBER_SHEETS = {
    'Vishal': 'vishal',
    'Kasthuri': 'kasthuri',
    'Kimaya': 'kimaya',
    'Son': 'son',
}

# section title keyword -> (collection id, field-keyword map)
# field-keyword map: canonical field key -> list of substrings to match
# (case-insensitive) against a header cell to claim that column.
SECTION_DEFS = [
    ('PERSONAL INFORMATION', 'profile', None),
    ('ACTIVE CONDITIONS', 'conditions', {
        'condition': ['condition', 'diagnosis'],
        'dateDiagnosed': ['date diagnosed'],
        'status': ['status'],
        'treatingDoctor': ['treating doctor'],
        'hospital': ['hospital', 'clinic'],
        'notes': ['notes'],
    }),
    ('CURRENT MEDICATIONS', 'medications', {
        'medication': ['medication', 'drug'],
        'dosage': ['dosage'],
        'frequency': ['frequency'],
        'prescribedBy': ['prescribed by'],
        'startDate': ['start date'],
        'endDate': ['end date'],
        'notes': ['notes'],
    }),
    ('ALLERGIES', 'allergies', {
        'allergen': ['allergen'],
        'type': ['type'],
        'reaction': ['reaction'],
        'severity': ['severity'],
        'diagnosedDate': ['diagnosed date'],
        'notes': ['notes'],
    }),
    ('SURGERIES', 'surgeries', {
        'procedure': ['procedure'],
        'date': ['date'],
        'hospital': ['hospital'],
        'surgeon': ['surgeon', 'doctor'],
        'notes': ['notes'],
    }),
    ('VACCINATIONS', 'vaccinations', {
        'vaccine': ['vaccine'],
        'date': ['date given', 'date'],
        'doseOrBooster': ['booster', 'batch', 'dose'],
        'hospital': ['hospital', 'clinic'],
        'notes': ['notes'],
    }),
    ('DOCTORS', 'doctors', {
        'specialty': ['speciality', 'specialty'],
        'name': ['doctor name', 'name'],
        'hospital': ['hospital', 'clinic'],
        'phone': ['phone'],
        'email': ['email'],
        'notes': ['notes'],
    }),
    # Includes suffixed variants like "CONSULTATION HISTORY — OPHTHALMOLOGY (...)"
    ('CONSULTATION HISTORY', 'consultations', {
        'date': ['date'],
        'doctor': ['doctor'],
        'specialty': ['speciality', 'specialty'],
        'hospital': ['hospital'],
        'reason': ['reason'],
        'diagnosis': ['diagnosis', 'findings'],
        'prescription': ['prescription', 'action'],
        'notes': ['notes'],
    }),
    ('LAB RESULTS', 'labResults', {
        'date': ['date'],
        'test': ['test', 'report'],
        'result': ['result'],
        'normalRange': ['normal range'],
        'status': ['status'],
        'lab': ['lab', 'clinic'],
        'orderedBy': ['ordered by'],
        'notes': ['notes'],
    }),
    ('HEALTH RECOMMENDATIONS', 'healthRecommendations', {
        'recommendation': ['recommendation', 'test'],
        'frequency': ['frequency'],
        'testsToInclude': ['tests to include'],
        'nextDue': ['next due'],
        'priority': ['priority'],
        'notes': ['notes'],
    }),
]


def match_section(title_text):
    upper = title_text.strip().upper()
    for keyword, collection_id, field_map in SECTION_DEFS:
        if keyword in upper:
            return collection_id, field_map
    return None, None


def cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if v is None:
        return ''
    return str(v).strip()


def map_headers(ws, header_row, field_map, max_col=8):
    """Return {col_index: field_key} for header_row using keyword matching."""
    col_to_field = {}
    used_keys = set()
    for c in range(1, max_col + 1):
        header = cell(ws, header_row, c).lower()
        if not header:
            continue
        for field_key, keywords in field_map.items():
            if field_key in used_keys:
                continue
            if any(kw in header for kw in keywords):
                col_to_field[c] = field_key
                used_keys.add(field_key)
                break
    return col_to_field


def parse_personal_info(ws, start_row, end_row, max_col=8):
    """Rows are label/value pairs, either one pair (A/B) or two pairs
    (A/B and C/D) per row, depending on the sheet."""
    data = {}
    label_to_key = {
        'full name': 'name',
        'date of birth': 'dob',
        'age': 'age',
        'blood group': 'bloodGroup',
        'primary doctor': 'primaryDoctor',
        'doctor phone': 'doctorPhone',
        'health insurance': 'healthInsurance',
        'policy number': 'policyNumber',
        'emergency contact': 'emergencyContact',
        'emergency phone': 'emergencyPhone',
        'birth details': 'birthDetails',
    }
    for r in range(start_row, end_row):
        pairs = [(1, 2), (3, 4)] if max_col >= 4 else [(1, 2)]
        for label_col, value_col in pairs:
            label = cell(ws, r, label_col).lower()
            value = cell(ws, r, value_col)
            key = label_to_key.get(label)
            if key and value:
                data[key] = value
    return data


def find_section_boundaries(ws):
    """Returns list of (title_text, start_data_row, end_row_exclusive)."""
    section_rows = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.startswith('  ') and v.strip():
            section_rows.append(r)
    boundaries = []
    for i, r in enumerate(section_rows):
        end = section_rows[i + 1] if i + 1 < len(section_rows) else ws.max_row + 1
        title = ws.cell(row=r, column=1).value.strip()
        boundaries.append((title, r, end))
    return boundaries


def parse_sheet(ws, member_id):
    """Returns (profile_dict, {collection_id: [record_dict, ...]}, skipped_titles)."""
    profile = {}
    records = {}
    skipped = []

    for title, start_row, end_row in find_section_boundaries(ws):
        collection_id, field_map = match_section(title)
        if collection_id is None:
            skipped.append(title)
            continue

        if collection_id == 'profile':
            profile.update(parse_personal_info(ws, start_row + 1, end_row))
            continue

        header_row = start_row + 1
        col_to_field = map_headers(ws, header_row, field_map)
        if not col_to_field:
            skipped.append(f'{title} (no headers matched)')
            continue

        rows = []
        for r in range(header_row + 1, end_row):
            row_data = {}
            for c, field_key in col_to_field.items():
                val = cell(ws, r, c)
                if val:
                    row_data[field_key] = val
            if row_data:
                rows.append(row_data)

        records.setdefault(collection_id, []).extend(rows)

    return profile, records, skipped


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('xlsx_path')
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)

    all_skipped = {}
    plan = {}  # member_id -> (profile, records)
    for sheet_name, member_id in MEMBER_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        profile, records, skipped = parse_sheet(wb[sheet_name], member_id)
        plan[member_id] = (profile, records)
        if skipped:
            all_skipped[member_id] = skipped

    print('=== Import plan ===')
    for member_id, (profile, records) in plan.items():
        print(f'\n{member_id}:')
        print(f'  profile fields: {len(profile)}')
        for collection_id, rows in records.items():
            print(f'  {collection_id}: {len(rows)} records')
    if all_skipped:
        print('\n=== Skipped sections (review manually against the xlsx) ===')
        for member_id, titles in all_skipped.items():
            for t in titles:
                print(f'  {member_id}: {t}')

    if args.dry_run:
        print('\n--dry-run: nothing written. Re-run without the flag to import.')
        return

    import firebase_admin
    from firebase_admin import credentials, firestore

    firebase_admin.initialize_app(credentials.ApplicationDefault())
    db = firestore.client()

    for member_id, (profile, records) in plan.items():
        db.collection('familyMembers').document(member_id).set(profile, merge=True)
        for collection_id, rows in records.items():
            col_ref = db.collection('familyMembers').document(member_id).collection(collection_id)
            for row in rows:
                col_ref.add(row)
        print(f'Imported {member_id}: {len(profile)} profile fields, '
              f'{sum(len(r) for r in records.values())} records across {len(records)} sections')

    print('\nDone.')


if __name__ == '__main__':
    sys.exit(main())
